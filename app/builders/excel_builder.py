"""Builder: construct Excel (XLSX) from tables step by step."""

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


class ExcelExportBuilder:
    """Build an XLSX file by adding sheets and tables."""

    def __init__(self) -> None:
        self._wb = Workbook()
        self._wb.remove(self._wb.active)
        self._current_sheet = None
        self._row_offset = 0

    def add_sheet(self, name: str) -> "ExcelExportBuilder":
        """Start a new sheet (max 31 chars for name)."""
        sheet_name = (name or "Sheet")[:31]
        self._current_sheet = self._wb.create_sheet(title=sheet_name)
        self._row_offset = 0
        return self

    def add_table(self, rows: list[list[Any]]) -> "ExcelExportBuilder":
        """Append a table (list of rows) to the current sheet. First row is header."""
        if not self._current_sheet:
            self.add_sheet("Sheet1")
        if not rows:
            return self
        df = pd.DataFrame(rows).replace({None: ""})
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                self._current_sheet.cell(
                    row=self._row_offset + r_idx,
                    column=c_idx,
                    value=value,
                )
        self._row_offset += len(df) + 2  # data + header + blank row
        return self

    def build(self) -> bytes:
        """Return XLSX file as bytes."""
        buffer = io.BytesIO()
        self._wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
