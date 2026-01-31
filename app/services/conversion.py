"""PDF to Excel conversion in memory. No persistent storage of PDFs."""

import io
import time
from typing import BinaryIO

import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from app.config import settings


class ConversionError(Exception):
    """Raised when PDF cannot be converted (no tables, corrupted, etc.)."""

    def __init__(self, message: str, code: str = "CONVERSION_ERROR"):
        self.message = message
        self.code = code
        super().__init__(message)


def validate_pdf(
    content: bytes,
    content_type: str | None,
    max_bytes: int | None = None,
    max_pages: int | None = None,
) -> None:
    """Validate PDF before processing. Raises ConversionError if invalid."""
    if content_type and content_type.lower() != "application/pdf":
        raise ConversionError("Only application/pdf is accepted.", "UNSUPPORTED_MIME")
    if max_bytes is None:
        max_bytes = settings.max_pdf_bytes
    if len(content) > max_bytes:
        raise ConversionError(
            f"File too large. Maximum size is {max_bytes // (1024 * 1024)} MB.",
            "FILE_TOO_LARGE",
        )
    if not content or len(content) < 100:
        raise ConversionError("File is empty or too small to be a valid PDF.", "PDF_CORRUPTED")
    try:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            num_pages = len(pdf.pages)
    except Exception as e:
        raise ConversionError("Unsupported or corrupted PDF.", "PDF_CORRUPTED") from e
    if max_pages is None:
        max_pages = settings.max_pdf_pages
    if num_pages > max_pages:
        raise ConversionError(
            f"Too many pages. Maximum is {max_pages}.",
            "PAGE_LIMIT_EXCEEDED",
        )


def pdf_to_xlsx_bytes(content: bytes, output_filename_base: str = "export") -> bytes:
    """
    Extract tables from PDF and return XLSX file as bytes.
    Uses pdfplumber for table extraction and openpyxl for Excel output.
    Does not store the PDF on disk.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        sheet_index = 0
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            if not tables:
                continue
            sheet_name = f"Page {page_num}"[:31]  # Excel sheet name max 31 chars
            ws = wb.create_sheet(title=sheet_name, index=sheet_index)
            sheet_index += 1
            row_offset = 0
            for table in tables:
                if not table:
                    continue
                df = pd.DataFrame(table)
                df = df.replace({None: ""})
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                    for c_idx, value in enumerate(row, start=1):
                        cell = ws.cell(row=row_offset + r_idx, column=c_idx, value=value)
                row_offset += len(df) + 2  # data + header + blank row between tables

    if sheet_index == 0:
        raise ConversionError("No table detected in PDF.", "NO_TABLE_DETECTED")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def convert_pdf_to_excel(
    file_content: bytes,
    filename: str,
    content_type: str | None = "application/pdf",
) -> tuple[bytes, float]:
    """
    Validate PDF, convert to XLSX, return (xlsx_bytes, duration_seconds).
    Raises ConversionError on validation or extraction failure.
    """
    validate_pdf(file_content, content_type)
    base = filename.rsplit(".", 1)[0] if "." in filename else filename
    start = time.perf_counter()
    xlsx_bytes = pdf_to_xlsx_bytes(file_content, output_filename_base=base)
    duration = time.perf_counter() - start
    return xlsx_bytes, duration
